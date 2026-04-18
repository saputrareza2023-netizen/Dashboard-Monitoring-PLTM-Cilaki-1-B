import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import numpy as np
from openpyxl import load_workbook
from datetime import timedelta, date
import requests
import json
import io

# ── Page Config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Monitoring PLTM Cilaki 1-B",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

  /* ── GLOBAL ── */
  html, body, [class*="css"], [class*="st-"], div, p, span, label {
    font-family: 'Inter', sans-serif !important;
    color: #1a1a2e !important;
  }

  /* ── BACKGROUND ── */
  .stApp, .main, [data-testid="stAppViewContainer"],
  [data-testid="stAppViewBlockContainer"], .block-container {
    background-color: #ffffff !important;
  }

  /* ── HEADER ── */
  header[data-testid="stHeader"] {
    background-color: #1c2333 !important;
    border-bottom: none !important;
  }
  header[data-testid="stHeader"] * { color: #ffffff !important; }

  /* ── SIDEBAR ── */
  section[data-testid="stSidebar"],
  section[data-testid="stSidebar"] > div,
  [data-testid="stSidebarContent"] {
    background-color: #1c2333 !important;
    border-right: none !important;
  }
  section[data-testid="stSidebar"] * {
    color: #e2e8f0 !important;
    background-color: transparent !important;
  }
  section[data-testid="stSidebar"] h1,
  section[data-testid="stSidebar"] h2,
  section[data-testid="stSidebar"] h3 {
    color: #ffffff !important;
  }
  section[data-testid="stSidebar"] .stButton > button {
    background-color: #3b7dd8 !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 4px !important;
    font-weight: 600 !important;
    font-size: 13px !important;
  }
  section[data-testid="stSidebar"] .stButton > button:hover {
    background-color: #2563c7 !important;
  }
  section[data-testid="stSidebar"] [data-testid="stFileUploader"] {
    background-color: #243048 !important;
    border: 1px dashed #4a6080 !important;
    border-radius: 6px !important;
  }

  /* ── TABS ── */
  .stTabs [data-baseweb="tab-list"] {
    background-color: #ffffff !important;
    border-bottom: 2px solid #e2e8f0 !important;
    border-radius: 0 !important;
    gap: 0 !important;
  }
  .stTabs [data-baseweb="tab"] {
    font-size: 13px !important;
    font-weight: 500 !important;
    color: #64748b !important;
    border-radius: 0 !important;
    background: transparent !important;
    padding: 10px 20px !important;
    border-bottom: 2px solid transparent !important;
    margin-bottom: -2px !important;
  }
  .stTabs [aria-selected="true"] {
    color: #3b7dd8 !important;
    border-bottom: 2px solid #3b7dd8 !important;
    background: transparent !important;
    font-weight: 600 !important;
  }
  .stTabs [data-baseweb="tab-panel"],
  .stTabs [data-baseweb="tab-panel"] > div {
    background-color: #ffffff !important;
  }

  /* ── BUTTONS ── */
  .stButton > button {
    border-radius: 4px !important;
    font-size: 12px !important;
    font-weight: 500 !important;
  }
  .stButton > button[kind="primary"] {
    background-color: #3b7dd8 !important;
    color: #ffffff !important;
    border: none !important;
  }
  .stButton > button[kind="secondary"] {
    background-color: #ffffff !important;
    color: #3b7dd8 !important;
    border: 1px solid #3b7dd8 !important;
  }

  /* ── INPUTS ── */
  .stSelectbox > div > div,
  .stMultiSelect > div > div,
  [data-baseweb="select"] > div {
    background-color: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 4px !important;
    color: #1a1a2e !important;
  }
  [data-baseweb="popover"] *, [data-baseweb="menu"] * {
    background-color: #ffffff !important;
    color: #1a1a2e !important;
  }
  [data-testid="stDateInput"] input {
    background-color: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 4px !important;
    color: #1a1a2e !important;
  }
  .stSlider [data-baseweb="slider"] div[role="slider"] {
    background-color: #3b7dd8 !important;
  }

  /* ── DATAFRAME ── */
  .stDataFrame, .stDataFrame thead tr th {
    background-color: #ffffff !important;
    color: #1a1a2e !important;
    font-size: 13px !important;
    border-color: #e2e8f0 !important;
  }
  .stDataFrame thead tr th {
    background-color: #f8fafc !important;
    font-weight: 600 !important;
    color: #475569 !important;
  }

  /* ── ALERTS ── */
  .stAlert {
    background-color: #eff6ff !important;
    border-left: 3px solid #3b7dd8 !important;
    border-radius: 4px !important;
    color: #1e40af !important;
  }

  /* ── METRIC CARDS ── */
  .metric-card {
    background: #ffffff !important;
    border: 1px solid #e2e8f0;
    border-top: 3px solid #3b7dd8;
    border-radius: 6px;
    padding: 16px 18px;
    text-align: left;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
  }
  .metric-card.warning { border-top-color: #f59e0b; }
  .metric-card.danger  { border-top-color: #ef4444; }
  .metric-card.success { border-top-color: #10b981; }

  .metric-label {
    font-size: 10px !important;
    font-weight: 600 !important;
    letter-spacing: 1px;
    text-transform: uppercase;
    color: #94a3b8 !important;
    margin-bottom: 8px;
  }
  .metric-value {
    font-size: 26px !important;
    color: #1a1a2e !important;
    font-weight: 700;
    line-height: 1;
  }
  .metric-sub {
    font-size: 11px !important;
    color: #94a3b8 !important;
    margin-top: 6px;
  }
  .metric-card.success .metric-value { color: #10b981 !important; }
  .metric-card.danger  .metric-value { color: #ef4444 !important; }
  .metric-card.warning .metric-value { color: #f59e0b !important; }

  /* ── SECTION HEADER ── */
  .section-header {
    font-size: 11px !important;
    font-weight: 600 !important;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    color: #64748b !important;
    border-bottom: 1px solid #e2e8f0;
    padding-bottom: 8px;
    margin-bottom: 16px;
    background: transparent !important;
    display: block;
  }

  /* ── TITLES ── */
  .dash-title {
    font-size: 20px !important;
    color: #1a1a2e !important;
    font-weight: 700;
    letter-spacing: -0.3px;
  }
  .dash-subtitle {
    font-size: 12px !important;
    color: #94a3b8 !important;
    letter-spacing: 0.5px;
    margin-top: 2px;
  }

  /* ── SCROLLBAR ── */
  ::-webkit-scrollbar { width: 6px; height: 6px; }
  ::-webkit-scrollbar-track { background: #f8fafc; }
  ::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 3px; }

  /* ── BLOCK CONTAINER ── */
  .block-container { padding-top: 1.5rem !important; max-width: 100% !important; }

  /* ── FILE UPLOADER ── */
  [data-testid="stFileUploaderDropzoneInstructions"] div:last-child {
    display: none !important;
  }
  [data-testid="stFileUploaderDropzone"] {
    padding: 12px !important;
    border-radius: 6px !important;
  }
</style>
""", unsafe_allow_html=True)


# ── Supabase REST API ─────────────────────────────────────────────────────────
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal"
}

def sb_select(table, params=""):
    url = f"{SUPABASE_URL}/rest/v1/{table}?{params}"
    r = requests.get(url, headers=HEADERS)
    if r.status_code == 200:
        data = r.json()
        if isinstance(data, list):
            return data
        return []
    else:
        st.error(f"Error ambil data: {r.status_code}")
        return []

def sb_upsert(table, data):
    url = f"{SUPABASE_URL}/rest/v1/{table}?on_conflict=tanggal,jam"
    headers = {**HEADERS, "Prefer": "resolution=merge-duplicates,return=minimal"}
    r = requests.post(url, headers=headers, data=json.dumps(data))
    if r.status_code not in [200, 201, 204]:
        st.error(f"Status: {r.status_code} | Error: {r.text}")
        return False
    return True

def sb_delete(table, eq_col, eq_val):
    url = f"{SUPABASE_URL}/rest/v1/{table}?{eq_col}=eq.{eq_val}"
    r = requests.delete(url, headers=HEADERS)
    return r.status_code in [200, 204]


# ── Helpers ───────────────────────────────────────────────────────────────────
def td_to_str(val):
    if isinstance(val, timedelta):
        h = int(val.total_seconds()) // 3600
        return f"{h:02d}:00"
    return str(val) if val is not None else ""

def num(v):
    if isinstance(v, (int, float)) and not (isinstance(v, float) and np.isnan(v)):
        return float(v)
    return None

DARK_BG  = "#ffffff"
GRID_COL = "#f1f5f9"
FONT_COL = "#1a1a2e"
LAYOUT   = dict(
    plot_bgcolor=DARK_BG, paper_bgcolor=DARK_BG,
    font=dict(color=FONT_COL, family="Barlow"),
    margin=dict(t=30, b=50, l=10, r=10)
)

def axis(title=""):
    return dict(title=title, gridcolor="#e2e8f0", color="#111111", zerolinecolor="#e2e8f0", tickfont=dict(family="Inter", size=12, color="#111111"), title_font=dict(color="#111111", size=12))

def kpi(col, label, value, sub, cls=""):
    col.markdown(
        f'<div class="metric-card {cls}">'
        f'<div class="metric-label">{label}</div>'
        f'<div class="metric-value">{value}</div>'
        f'<div class="metric-sub">{sub}</div>'
        f'</div>', unsafe_allow_html=True)


# ── Parse Excel harian ────────────────────────────────────────────────────────
BULAN_MAP = {
    "januari":1,"februari":2,"maret":3,"april":4,
    "mei":5,"juni":6,"juli":7,"agustus":8,
    "september":9,"oktober":10,"november":11,"desember":12,
    "january":1,"february":2,"march":3,"may":5,
    "june":6,"july":7,"august":8,"october":10,"december":12
}

def detect_date_from_filename(filename):
    import re
    name = filename.lower().replace("_"," ").replace("-"," ")
    tahun, bulan = None, None
    years = re.findall(r'\b(20\d{2})\b', name)
    if years:
        tahun = int(years[-1])
    for bln_name, bln_num in BULAN_MAP.items():
        if bln_name in name:
            bulan = bln_num
            break
    if not bulan:
        nums = re.findall(r'\b(0?[1-9]|1[0-2])\b', name)
        if nums:
            bulan = int(nums[0])
    return tahun, bulan

def parse_excel_harian(file_bytes, filename, tanggal_override=None):
    from datetime import date as date_cls
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_days = sorted([int(n) for n in wb.sheetnames if n.isdigit() and 1 <= int(n) <= 31])
    tahun, bulan = detect_date_from_filename(filename)
    data = []

    def build_row(r, tgl):
        jam = td_to_str(r[0])
        if not jam or jam == "None":
            return None
        if all(r[i] is None for i in [2, 86, 89] if i < len(r)):
            return None
        return {
            "tanggal": str(tgl), "jam": jam,
            "tg1_mw": num(r[2]), "tg1_pf": num(r[3]),
            "tg2_mw": num(r[30]) if len(r)>30 else None,
            "tg3_mw": num(r[58]) if len(r)>58 else None,
            "total_mw": num(r[86]) if len(r)>86 else None,
            "total_pf": num(r[87]) if len(r)>87 else None,
            "total_mvar": num(r[88]) if len(r)>88 else None,
            "volt_r": num(r[89]) if len(r)>89 else None,
            "volt_s": num(r[90]) if len(r)>90 else None,
            "volt_t": num(r[91]) if len(r)>91 else None,
        }

    if sheet_days:
        for day in sheet_days:
            ws = wb[str(day)]
            rows = list(ws.iter_rows(min_row=1, max_row=32, values_only=True))
            try:
                tgl = date_cls(tahun, bulan, day) if tahun and bulan else (tanggal_override or date.today())
            except:
                tgl = tanggal_override or date.today()
            for r in rows[3:27]:
                if r[0] is None: continue
                row = build_row(r, tgl)
                if row: data.append(row)
    else:
        target_sheet = next((n for n in wb.sheetnames if n.lower() not in ["rekap","summary","rekapitulasi"]), wb.sheetnames[0])
        ws = wb[target_sheet]
        rows = list(ws.iter_rows(min_row=1, max_row=32, values_only=True))
        tgl = tanggal_override or date.today()
        for r in rows[3:27]:
            if r[0] is None: continue
            row = build_row(r, tgl)
            if row: data.append(row)
    return data


# ── Load data ─────────────────────────────────────────────────────────────────
def load_data():
    rows = sb_select("data_harian", "order=tanggal.asc,jam.asc&limit=50000")
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df["tanggal"] = pd.to_datetime(df["tanggal"]).dt.date
    return df

def load_data_debug():
    """Load data with debug info."""
    rows = sb_select("data_harian", "select=tanggal&order=tanggal.asc&limit=50000")
    return rows

def hitung_summary(df):
    if df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    cols = ["total_mw","tg1_mw","tg2_mw","tg3_mw","total_pf","total_mvar","volt_r","volt_s","volt_t"]
    df_max = df.groupby("tanggal")[cols].max().reset_index()
    df_avg = df.groupby("tanggal")[cols].mean().reset_index()
    df_min = df.groupby("tanggal")[cols].min().reset_index()
    return df_max, df_avg, df_min


# ════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("<div style='font-size:13px;font-weight:700;color:#ffffff;letter-spacing:0.5px'>⚡ PLTM CILAKI 1-B</div>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("**📂 Upload Data Harian**")
    st.markdown("""<style>
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzoneInstructions"]>div>span{display:none!important}
    section[data-testid="stSidebar"] [data-testid="stFileUploader"] label{display:none!important}
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] small{display:none!important}
    </style>""", unsafe_allow_html=True)
    uploaded = st.file_uploader("Upload", type=["xlsx","xls"], label_visibility="collapsed")
    st.caption("📅 Tanggal otomatis dari nama file")

    if uploaded and st.button("💾 Simpan ke Database", type="primary", use_container_width=True):
        with st.spinner("Memproses dan menyimpan data..."):
            file_bytes = uploaded.read()
            data_list  = parse_excel_harian(file_bytes, uploaded.name)
            if data_list:
                ok = sb_upsert("data_harian", data_list)
                if ok:
                    st.success(f"✅ {len(data_list)} data jam berhasil disimpan!")
                    st.cache_data.clear()
                    st.rerun()
                else:
                    st.error("❌ Gagal menyimpan ke database.")
            else:
                st.error("❌ Tidak ada data yang terbaca dari file.")

    st.markdown("---")
    st.markdown("**⚙️ Threshold Tegangan (kV)**")
    volt_min   = st.slider("Tegangan Minimum",  18.0, 21.0, 20.0, 0.1)
    volt_max_v = st.slider("Tegangan Maksimum", 20.0, 23.0, 21.5, 0.1)
    st.markdown("**⚙️ Batas Beban Normal**")
    beban_max  = st.slider("Beban Maks (MW)", 1.0, 5.0, 2.8, 0.1)
    st.markdown("---")
    st.markdown(
        "<div style='color:#94a3b8;font-size:11px;line-height:1.7'>"
        "📊 Dashboard Monitoring<br>PLTM Cilaki 1-B<br>"
        "Sistem 20 kV · TG1·TG2·TG3"
        "</div>", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
#  HEADER
# ════════════════════════════════════════════════════════════════════════════
st.markdown(
    '<div class="dash-title">PLTM Cilaki 1-B</div>'
    '<div class="dash-subtitle">Monitoring Profil Tegangan &amp; Beban · Real Time</div>',
    unsafe_allow_html=True)
st.markdown("")

# ── Load data ─────────────────────────────────────────────────────────────────
df = load_data()

# Debug
if not df.empty:
    st.sidebar.caption(f"📊 Total data: {len(df)} baris | {df['tanggal'].nunique()} hari | {df['tanggal'].min()} s/d {df['tanggal'].max()}")

if df.empty:
    st.info("📭 Belum ada data. Silakan upload file Excel harian di sidebar.")
    st.markdown("""
    ### 📋 Cara Upload Data Harian
    1. Pilih **tanggal** data di sidebar
    2. Klik **Browse files** dan pilih file Excel harian
    3. Klik tombol **Simpan ke Database**
    4. Data otomatis tersimpan dan dashboard terupdate
    """)
    st.stop()

df_max, df_avg, df_min = hitung_summary(df)
df_max["tgl_str"] = df_max["tanggal"].astype(str)
df_avg["tgl_str"] = df_avg["tanggal"].astype(str)
df_min["tgl_str"] = df_min["tanggal"].astype(str)


# ── Tabs ──────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊  Ringkasan",
    "📈  Profil Beban",
    "⚡  Profil Tegangan",
    "🗂️  Data Lengkap",
    "📉  Month on Month",
])


# ════════════════════════════════════════════════════════════════════════════
#  TAB 1 — RINGKASAN
# ════════════════════════════════════════════════════════════════════════════
with tab1:
    st.markdown('<div class="section-header">STATISTIK KESELURUHAN</div>', unsafe_allow_html=True)

    tot_days = df["tanggal"].nunique()
    tot_max  = df_max["total_mw"].max()
    tot_avg  = df_avg["total_mw"].mean()
    vR_avg   = df_avg["volt_r"].mean()
    pf_avg   = df_avg["total_pf"].mean()
    last_tgl = df["tanggal"].max()

    k1,k2,k3,k4,k5 = st.columns(5)
    kpi(k1, "Data Tersimpan",     f"{tot_days} Hari",   f"s/d {last_tgl}")
    kpi(k2, "Beban Tertinggi",    f"{tot_max:.2f} MW",  "sepanjang periode")
    kpi(k3, "Beban Rata-rata",    f"{tot_avg:.2f} MW",  "rata-rata harian")
    kpi(k4, "Tegangan Rata-rata", f"{vR_avg:.3f} kV",   "phasa R",
        "success" if volt_min <= vR_avg <= volt_max_v else "warning")
    kpi(k5, "Power Factor",       f"{pf_avg:.4f}",      "rata-rata",
        "success" if pf_avg >= 0.95 else "warning")

    st.markdown("")

    # Tren beban dengan filter periode
    st.markdown('<div class="section-header">TREN BEBAN HARIAN (MW)</div>', unsafe_allow_html=True)

    # Tombol filter periode
    import datetime
    today     = pd.Timestamp(df["tanggal"].max())
    periode_opt = {"1D":1,"7D":7,"1M":30,"3M":90,"6M":180,"1Y":365,"ALL":99999}
    sel_col = st.columns(len(periode_opt))
    if "periode_sel" not in st.session_state:
        st.session_state["periode_sel"] = "ALL"
    for i,(label,_) in enumerate(periode_opt.items()):
        if sel_col[i].button(label, key=f"btn_{label}",
            use_container_width=True,
            type="primary" if st.session_state["periode_sel"]==label else "secondary"):
            st.session_state["periode_sel"] = label

    hari_filter = periode_opt[st.session_state["periode_sel"]]
    cutoff      = today - pd.Timedelta(days=hari_filter)

    df_max_f = df_max[pd.to_datetime(df_max["tanggal"]) >= cutoff].copy()
    df_avg_f = df_avg[pd.to_datetime(df_avg["tanggal"]) >= cutoff].copy()
    df_min_f = df_min[pd.to_datetime(df_min["tanggal"]) >= cutoff].copy()

    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=df_avg_f["tgl_str"],
        y=df_avg_f["total_mw"],
        mode="lines",
        name="Beban (MW)",
        line=dict(color="#3b7dd8", width=2.5),
        fill="tozeroy",
        fillcolor="rgba(59,125,216,0.08)",
        hovertemplate="<b>%{x}</b><br>Beban: %{y:.2f} MW<extra></extra>",
    ))

    fig.add_hline(y=beban_max, line_dash="dash", line_color="#f59e0b",
                  annotation_text=f"Batas {beban_max} MW",
                  annotation_font_color="#f59e0b")

    fig.update_layout(**LAYOUT, height=380,
        xaxis=dict(**axis("Tanggal"), tickangle=-45),
        yaxis=dict(**axis("Total Beban (MW)"), rangemode="tozero"),
        showlegend=False,
        hovermode="x unified",
        hoverlabel=dict(
            bgcolor="#1a1a2e",
            font_color="#ffffff",
            font_size=13,
            bordercolor="#3b7dd8",
        ))
    st.plotly_chart(fig, use_container_width=True)

    # Kontribusi per unit
    st.markdown('<div class="section-header">KONTRIBUSI BEBAN PER UNIT (MW)</div>', unsafe_allow_html=True)
    fig2 = go.Figure()
    fig2.add_trace(go.Bar(x=df_avg["tgl_str"], y=df_avg["tg1_mw"].fillna(0), name="TG1", marker_color="#3b7dd8"))
    fig2.add_trace(go.Bar(x=df_avg["tgl_str"], y=df_avg["tg2_mw"].fillna(0), name="TG2", marker_color="#2563c7"))
    fig2.add_trace(go.Bar(x=df_avg["tgl_str"], y=df_avg["tg3_mw"].fillna(0), name="TG3", marker_color="#10b981"))
    fig2.update_layout(**LAYOUT, barmode="stack", height=300,
        xaxis=dict(**axis("Tanggal"), tickangle=-45),
        yaxis=axis("Beban (MW)"),
        legend=dict(bgcolor="#ffffff", bordercolor="#e2e8f0", font=dict(color="#64748b", family="Inter", size=11)))
    st.plotly_chart(fig2, use_container_width=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="section-header">POWER FACTOR HARIAN</div>', unsafe_allow_html=True)
        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(x=df_avg["tgl_str"], y=df_avg["total_pf"],
            mode="lines+markers", name="PF Rata-rata",
            line=dict(color="#f59e0b",width=2), marker=dict(size=5)))
        fig3.add_hline(y=0.95, line_dash="dot", line_color="#ff4444",
                       annotation_text="Min 0.95", annotation_font_color="#ff4444")
        fig3.update_layout(**LAYOUT, height=260,
            xaxis=dict(**axis(), tickangle=-45),
            yaxis=dict(**axis("Power Factor"), range=[0.93,1.0]))
        st.plotly_chart(fig3, use_container_width=True)

    with c2:
        st.markdown('<div class="section-header">DAYA REAKTIF Q (MVAr)</div>', unsafe_allow_html=True)
        fig4 = go.Figure()
        fig4.add_trace(go.Bar(x=df_avg["tgl_str"], y=df_avg["total_mvar"],
            name="Q Rata-rata", marker_color="#8b5cf6"))
        fig4.add_trace(go.Scatter(x=df_max["tgl_str"], y=df_max["total_mvar"],
            mode="lines", name="Q Max", line=dict(color="#ef4444",dash="dot",width=2)))
        fig4.update_layout(**LAYOUT, height=260,
            xaxis=dict(**axis(), tickangle=-45),
            yaxis=axis("Q (MVAr)"))
        st.plotly_chart(fig4, use_container_width=True)


# ════════════════════════════════════════════════════════════════════════════
#  TAB 2 — PROFIL BEBAN HARIAN
# ════════════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown('<div class="section-header">PROFIL BEBAN PER JAM</div>', unsafe_allow_html=True)

    tanggal_list = sorted(df["tanggal"].unique(), reverse=True)
    col_a, col_b = st.columns([1,3])

    with col_a:
        pilih_tgl = st.selectbox(
            "Pilih Tanggal", options=tanggal_list,
            format_func=lambda x: str(x))
        tampil_unit = st.multiselect(
            "Unit", ["TG1","TG2","TG3","Total"],
            default=["TG1","TG2","TG3","Total"])

    df_day = df[df["tanggal"] == pilih_tgl].copy()

    with col_b:
        d_max = df_max[df_max["tanggal"] == pilih_tgl]
        d_avg = df_avg[df_avg["tanggal"] == pilih_tgl]
        d_min = df_min[df_min["tanggal"] == pilih_tgl]
        k1,k2,k3,k4 = st.columns(4)
        if not d_max.empty:
            kpi(k1, "Beban Max", f"{d_max['total_mw'].values[0]:.2f} MW", str(pilih_tgl))
            kpi(k2, "Beban Avg", f"{d_avg['total_mw'].values[0]:.2f} MW" if not d_avg.empty else "-", "")
            kpi(k3, "Beban Min", f"{d_min['total_mw'].values[0]:.2f} MW" if not d_min.empty else "-", "")
            pf_v = d_avg['total_pf'].values[0] if not d_avg.empty else 0
            kpi(k4, "PF Rata-rata", f"{pf_v:.4f}", "",
                "success" if pf_v >= 0.95 else "warning")

    unit_colors = {"TG1":"#3b7dd8","TG2":"#2563c7","TG3":"#10b981","Total":"#f59e0b"}
    unit_cols   = {"TG1":"tg1_mw","TG2":"tg2_mw","TG3":"tg3_mw","Total":"total_mw"}

    fig = go.Figure()
    for unit in tampil_unit:
        col = unit_cols[unit]
        mask = df_day[col].notna() & (df_day[col] > 0)
        fig.add_trace(go.Scatter(
            x=df_day.loc[mask,"jam"], y=df_day.loc[mask,col],
            mode="lines+markers", name=unit,
            line=dict(color=unit_colors[unit],width=2.5), marker=dict(size=6)))
    fig.add_hline(y=beban_max, line_dash="dash", line_color="#ff4444",
                  annotation_text=f"Batas {beban_max} MW", annotation_font_color="#ff4444")
    fig.update_layout(**LAYOUT, height=380,
        title=f"Profil Beban — {pilih_tgl}",
        title_font=dict(color="#e0f0ff",size=14),
        xaxis=dict(**axis("Jam"), tickangle=-45),
        yaxis=axis("Beban (MW)"),
        legend=dict(bgcolor="#ffffff", bordercolor="#e2e8f0", font=dict(color="#64748b", family="Inter", size=11)))
    st.plotly_chart(fig, use_container_width=True)

    st.markdown('<div class="section-header">DATA PER JAM</div>', unsafe_allow_html=True)
    show_cols = ["jam","tg1_mw","tg2_mw","tg3_mw","total_mw","total_pf","total_mvar"]
    df_show   = df_day[show_cols].rename(columns={
        "jam":"Jam","tg1_mw":"TG1 (MW)","tg2_mw":"TG2 (MW)","tg3_mw":"TG3 (MW)",
        "total_mw":"Total (MW)","total_pf":"PF","total_mvar":"Q (MVAr)"})

    def hl(row):
        mw = row.get("Total (MW)", None)
        if mw and mw >= beban_max:
            return ["background-color:#2a0000;color:#ffcccc"] * len(row)
        return [""] * len(row)

    st.dataframe(df_show.style.apply(hl, axis=1).format(
        {"TG1 (MW)":"{:.2f}","TG2 (MW)":"{:.2f}","TG3 (MW)":"{:.2f}",
         "Total (MW)":"{:.2f}","PF":"{:.4f}","Q (MVAr)":"{:.3f}"}, na_rep="-"),
        use_container_width=True, hide_index=True, height=380)


# ════════════════════════════════════════════════════════════════════════════
#  TAB 3 — PROFIL TEGANGAN
# ════════════════════════════════════════════════════════════════════════════
with tab3:
    st.markdown('<div class="section-header">TREN TEGANGAN (kV)</div>', unsafe_allow_html=True)

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df_max["tgl_str"], y=df_max["volt_r"],
        mode="lines+markers", name="Max R",
        line=dict(color="#ef4444",width=2), marker=dict(size=5)))
    fig.add_trace(go.Scatter(x=df_avg["tgl_str"], y=df_avg["volt_r"],
        mode="lines+markers", name="Rata-rata R",
        line=dict(color="#3b7dd8",width=2.5), marker=dict(size=5)))
    fig.add_trace(go.Scatter(x=df_min["tgl_str"], y=df_min["volt_r"],
        mode="lines+markers", name="Min R",
        line=dict(color="#10b981",width=2,dash="dot"), marker=dict(size=4)))
    fig.add_hrect(y0=volt_min, y1=volt_max_v,
        fillcolor="rgba(59,125,216,0.05)", line_width=0,
        annotation_text=f"Normal {volt_min}–{volt_max_v} kV",
        annotation_font_color="#4a7fa5", annotation_position="top left")
    fig.add_hline(y=volt_max_v, line_dash="dot", line_color="#f59e0b", line_width=1)
    fig.add_hline(y=volt_min,   line_dash="dot", line_color="#ff4444", line_width=1)
    fig.update_layout(**LAYOUT, height=360,
        xaxis=dict(**axis("Tanggal"), tickangle=-45),
        yaxis=dict(**axis("Tegangan (kV)"), range=[volt_min-0.5, volt_max_v+0.3]),
        legend=dict(bgcolor="#ffffff", bordercolor="#e2e8f0", font=dict(color="#64748b", family="Inter", size=11)))
    st.plotly_chart(fig, use_container_width=True)

    st.markdown('<div class="section-header">PROFIL TEGANGAN PER JAM</div>', unsafe_allow_html=True)
    pilih_tgl_v = st.selectbox(
        "Pilih Tanggal", options=tanggal_list,
        format_func=lambda x: str(x), key="volt_day")
    df_dv = df[df["tanggal"] == pilih_tgl_v].copy()

    fig2 = go.Figure()
    ph = {"volt_r":"#ef4444","volt_s":"#f59e0b","volt_t":"#10b981"}
    pn = {"volt_r":"Phasa R","volt_s":"Phasa S","volt_t":"Phasa T"}
    for col_name, color in ph.items():
        mask = df_dv[col_name].notna()
        fig2.add_trace(go.Scatter(
            x=df_dv.loc[mask,"jam"], y=df_dv.loc[mask,col_name],
            mode="lines+markers", name=pn[col_name],
            line=dict(color=color,width=2.5), marker=dict(size=5)))
    fig2.add_hrect(y0=volt_min, y1=volt_max_v, fillcolor="rgba(59,125,216,0.05)", line_width=0)
    fig2.add_hline(y=volt_max_v, line_dash="dot", line_color="#f59e0b", line_width=1)
    fig2.add_hline(y=volt_min,   line_dash="dot", line_color="#ff4444", line_width=1)
    fig2.update_layout(**LAYOUT, height=340,
        title=f"Tegangan R/S/T — {pilih_tgl_v}",
        title_font=dict(color="#e0f0ff",size=14),
        xaxis=dict(**axis("Jam"), tickangle=-45),
        yaxis=dict(**axis("Tegangan (kV)"), range=[volt_min-1, volt_max_v+0.5]),
        legend=dict(bgcolor="#ffffff", bordercolor="#e2e8f0", font=dict(color="#64748b", family="Inter", size=11)))
    st.plotly_chart(fig2, use_container_width=True)

    mask_low  = df["volt_r"].notna() & (df["volt_r"] < volt_min)
    mask_high = df["volt_r"].notna() & (df["volt_r"] > volt_max_v)
    if (mask_low | mask_high).any():
        st.markdown('<div class="section-header">⚠️ PELANGGARAN BATAS TEGANGAN</div>', unsafe_allow_html=True)
        viol = df[mask_low | mask_high][["tanggal","jam","volt_r","volt_s","volt_t"]].copy()
        viol["Status"] = viol["volt_r"].apply(
            lambda v: "UNDERVOLTAGE" if v < volt_min else "OVERVOLTAGE")
        st.dataframe(viol, use_container_width=True, hide_index=True)
    else:
        st.success("✅ Tidak ada pelanggaran batas tegangan pada periode ini.")


# ════════════════════════════════════════════════════════════════════════════
#  TAB 4 — DATA LENGKAP
# ════════════════════════════════════════════════════════════════════════════
with tab4:
    st.markdown('<div class="section-header">DATA LENGKAP</div>', unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        tgl_dari = st.date_input("Dari Tanggal", value=df["tanggal"].min())
    with c2:
        tgl_ke   = st.date_input("Sampai Tanggal", value=df["tanggal"].max())

    df_filtered = df[(df["tanggal"] >= tgl_dari) & (df["tanggal"] <= tgl_ke)].copy()
    show_c = ["tanggal","jam","tg1_mw","tg2_mw","tg3_mw",
              "total_mw","total_pf","total_mvar","volt_r","volt_s","volt_t"]
    fmt = {c:"{:.3f}" for c in show_c if c not in ["tanggal","jam"]}
    fmt["total_pf"] = "{:.4f}"

    st.dataframe(df_filtered[show_c].style.format(fmt, na_rep="-"),
        use_container_width=True, hide_index=True, height=500)
    st.markdown(f"**Total: {len(df_filtered)} baris data**")

    st.markdown("")
    ex1, ex2 = st.columns(2)
    with ex1:
        csv_out = df_filtered[show_c].to_csv(index=False).encode()
        st.download_button("📥 Export CSV", data=csv_out,
            file_name=f"data_pltm_{tgl_dari}_{tgl_ke}.csv",
            mime="text/csv", use_container_width=True)
    with ex2:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            df_filtered[show_c].to_excel(w, index=False, sheet_name="Data")
        st.download_button("📥 Export Excel", data=buf.getvalue(),
            file_name=f"data_pltm_{tgl_dari}_{tgl_ke}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

    st.markdown("---")
    st.markdown('<div class="section-header">🗑️ HAPUS DATA</div>', unsafe_allow_html=True)
    st.warning("⚠️ Hati-hati! Data yang dihapus tidak bisa dikembalikan.")
    tgl_hapus = st.date_input("Pilih tanggal yang ingin dihapus")
    if st.button("🗑️ Hapus Data Tanggal Ini", type="secondary"):
        ok = sb_delete("data_harian", "tanggal", str(tgl_hapus))
        if ok:
            st.success(f"✅ Data tanggal {tgl_hapus} berhasil dihapus!")
            st.cache_data.clear()
            st.rerun()
        else:
            st.error("❌ Gagal menghapus data.")


# ════════════════════════════════════════════════════════════════════════════
#  TAB 5 — MONTH ON MONTH
# ════════════════════════════════════════════════════════════════════════════
with tab5:
    st.markdown('<div class="section-header">PERBANDINGAN MONTH ON MONTH</div>', unsafe_allow_html=True)

    # Ambil daftar bulan & tahun yang tersedia
    df["tahun"]  = pd.to_datetime(df["tanggal"]).dt.year
    df["bulan"]  = pd.to_datetime(df["tanggal"]).dt.month
    df["hari"]   = pd.to_datetime(df["tanggal"]).dt.day

    tahun_list = sorted(df["tahun"].unique())
    bulan_list = sorted(df["bulan"].unique())
    bulan_nama = {1:"Januari",2:"Februari",3:"Maret",4:"April",5:"Mei",6:"Juni",
                  7:"Juli",8:"Agustus",9:"September",10:"Oktober",11:"November",12:"Desember"}

    if len(tahun_list) < 2:
        st.warning("⚠️ Perlu minimal data dari **2 tahun berbeda** untuk perbandingan MoM. Silakan upload data tahun lainnya.")
    else:
        # Pilihan filter
        c1, c2, c3 = st.columns(3)
        with c1:
            pilih_bulan = st.selectbox(
                "Pilih Bulan",
                options=bulan_list,
                format_func=lambda x: bulan_nama.get(x, str(x))
            )
        with c2:
            tahun_a = st.selectbox("Tahun Pertama", options=tahun_list, index=0)
        with c3:
            tahun_b = st.selectbox("Tahun Kedua",   options=tahun_list, index=len(tahun_list)-1)

        # Filter data
        df_a = df[(df["bulan"] == pilih_bulan) & (df["tahun"] == tahun_a)].copy()
        df_b = df[(df["bulan"] == pilih_bulan) & (df["tahun"] == tahun_b)].copy()

        # Hitung rata-rata per hari
        cols_agg = ["hari","total_mw","volt_r","volt_s","volt_t","total_pf","total_mvar"]
        avg_a = df_a.groupby("hari")[["total_mw","volt_r","volt_s","volt_t","total_pf","total_mvar"]].mean().reset_index()
        avg_b = df_b.groupby("hari")[["total_mw","volt_r","volt_s","volt_t","total_pf","total_mvar"]].mean().reset_index()

        nama_a = f"{bulan_nama.get(pilih_bulan,'')} {tahun_a}"
        nama_b = f"{bulan_nama.get(pilih_bulan,'')} {tahun_b}"

        if avg_a.empty and avg_b.empty:
            st.warning("Tidak ada data untuk bulan dan tahun yang dipilih.")
        else:
            # ── KPI perbandingan ──────────────────────────────────────────
            st.markdown("")
            k1,k2,k3,k4 = st.columns(4)
            avg_mw_a = avg_a["total_mw"].mean() if not avg_a.empty else 0
            avg_mw_b = avg_b["total_mw"].mean() if not avg_b.empty else 0
            delta_mw = avg_mw_b - avg_mw_a
            delta_pct = (delta_mw / avg_mw_a * 100) if avg_mw_a > 0 else 0

            avg_v_a = avg_a["volt_r"].mean() if not avg_a.empty else 0
            avg_v_b = avg_b["volt_r"].mean() if not avg_b.empty else 0
            delta_v = avg_v_b - avg_v_a

            kpi(k1, f"Avg Beban {tahun_a}", f"{avg_mw_a:.2f} MW", nama_a)
            kpi(k2, f"Avg Beban {tahun_b}", f"{avg_mw_b:.2f} MW", nama_b,
                "success" if delta_mw >= 0 else "danger")
            kpi(k3, "Selisih Beban",
                f"{'↑' if delta_mw>=0 else '↓'} {abs(delta_mw):.2f} MW",
                f"{delta_pct:+.1f}%",
                "success" if delta_mw >= 0 else "danger")
            kpi(k4, "Selisih Tegangan",
                f"{'↑' if delta_v>=0 else '↓'} {abs(delta_v):.3f} kV",
                f"{tahun_a} vs {tahun_b}",
                "success" if abs(delta_v) < 0.5 else "warning")

            st.markdown("")

            # ── Grafik Beban MoM ──────────────────────────────────────────
            st.markdown('<div class="section-header">PERBANDINGAN BEBAN TOTAL (MW)</div>', unsafe_allow_html=True)
            fig_mom = go.Figure()
            if not avg_a.empty:
                fig_mom.add_trace(go.Scatter(
                    x=avg_a["hari"], y=avg_a["total_mw"],
                    mode="lines+markers", name=nama_a,
                    line=dict(color="#3b7dd8", width=2.5),
                    marker=dict(size=6),
                ))
            if not avg_b.empty:
                fig_mom.add_trace(go.Scatter(
                    x=avg_b["hari"], y=avg_b["total_mw"],
                    mode="lines+markers", name=nama_b,
                    line=dict(color="#ef4444", width=2.5),
                    marker=dict(size=6),
                ))
            # Area diff
            if not avg_a.empty and not avg_b.empty:
                merged = pd.merge(avg_a[["hari","total_mw"]], avg_b[["hari","total_mw"]],
                                  on="hari", suffixes=("_a","_b"))
                fig_mom.add_trace(go.Scatter(
                    x=merged["hari"], y=merged["total_mw_b"],
                    fill="tonexty", fillcolor="rgba(255,107,107,0.1)",
                    line=dict(width=0), showlegend=False, name="Selisih"
                ))

            fig_mom.add_hline(y=beban_max, line_dash="dash", line_color="#f59e0b",
                              annotation_text=f"Batas {beban_max} MW",
                              annotation_font_color="#f59e0b")
            fig_mom.update_layout(**LAYOUT, height=380,
                xaxis=dict(**axis("Hari ke-"), dtick=1, range=[0.5, 31.5]),
                yaxis=axis("Beban (MW)"),
                legend=dict(bgcolor="#ffffff", bordercolor="#e2e8f0", font=dict(color="#64748b", family="Inter", size=11)),
                hovermode="x unified")
            st.plotly_chart(fig_mom, use_container_width=True)

            # ── Grafik Tegangan MoM ───────────────────────────────────────
            st.markdown('<div class="section-header">PERBANDINGAN TEGANGAN PHASA R (kV)</div>', unsafe_allow_html=True)
            fig_volt = go.Figure()
            if not avg_a.empty:
                fig_volt.add_trace(go.Scatter(
                    x=avg_a["hari"], y=avg_a["volt_r"],
                    mode="lines+markers", name=nama_a,
                    line=dict(color="#3b7dd8", width=2.5),
                    marker=dict(size=6),
                ))
            if not avg_b.empty:
                fig_volt.add_trace(go.Scatter(
                    x=avg_b["hari"], y=avg_b["volt_r"],
                    mode="lines+markers", name=nama_b,
                    line=dict(color="#ef4444", width=2.5),
                    marker=dict(size=6),
                ))
            fig_volt.add_hrect(y0=volt_min, y1=volt_max_v,
                fillcolor="rgba(59,125,216,0.05)", line_width=0,
                annotation_text=f"Normal {volt_min}–{volt_max_v} kV",
                annotation_font_color="#4a7fa5", annotation_position="top left")
            fig_volt.add_hline(y=volt_max_v, line_dash="dot", line_color="#f59e0b", line_width=1)
            fig_volt.add_hline(y=volt_min,   line_dash="dot", line_color="#ff4444", line_width=1)
            fig_volt.update_layout(**LAYOUT, height=340,
                xaxis=dict(**axis("Hari ke-"), dtick=1, range=[0.5, 31.5]),
                yaxis=dict(**axis("Tegangan (kV)"), range=[volt_min-0.5, volt_max_v+0.3]),
                legend=dict(bgcolor="#ffffff", bordercolor="#e2e8f0", font=dict(color="#64748b", family="Inter", size=11)),
                hovermode="x unified")
            st.plotly_chart(fig_volt, use_container_width=True)

            # ── Tabel perbandingan ────────────────────────────────────────
            st.markdown('<div class="section-header">TABEL PERBANDINGAN PER HARI</div>', unsafe_allow_html=True)
            if not avg_a.empty and not avg_b.empty:
                tbl = pd.merge(
                    avg_a[["hari","total_mw","volt_r"]].rename(columns={"total_mw":f"MW {tahun_a}","volt_r":f"V_R {tahun_a}"}),
                    avg_b[["hari","total_mw","volt_r"]].rename(columns={"total_mw":f"MW {tahun_b}","volt_r":f"V_R {tahun_b}"}),
                    on="hari", how="outer"
                ).sort_values("hari")
                tbl["Δ MW"]  = tbl[f"MW {tahun_b}"]  - tbl[f"MW {tahun_a}"]
                tbl["Δ V_R"] = tbl[f"V_R {tahun_b}"] - tbl[f"V_R {tahun_a}"]
                tbl["Hari"]  = tbl["hari"].astype(int)

                def color_delta(val):
                    if pd.isna(val): return ""
                    return "color:#00e676" if val >= 0 else "color:#ff4444"

                fmt = {f"MW {tahun_a}":"{:.2f}", f"MW {tahun_b}":"{:.2f}",
                       f"V_R {tahun_a}":"{:.3f}", f"V_R {tahun_b}":"{:.3f}",
                       "Δ MW":"{:+.2f}", "Δ V_R":"{:+.3f}"}
                show = ["Hari", f"MW {tahun_a}", f"MW {tahun_b}", "Δ MW", f"V_R {tahun_a}", f"V_R {tahun_b}", "Δ V_R"]
                st.dataframe(
                    tbl[show].style.applymap(color_delta, subset=["Δ MW","Δ V_R"]).format(fmt, na_rep="-"),
                    use_container_width=True, hide_index=True, height=400
                )
